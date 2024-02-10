from openpyxl import load_workbook


class Read_data:

    def get_test_data():
        workbook = load_workbook("C:\\Users\\Administrator\\Desktop\\Book1.xlsx")
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)
        return data
